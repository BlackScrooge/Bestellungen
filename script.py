from openpyxl import Workbook, load_workbook

wb = load_workbook("Bestellungen.xlsx")
ws = wb["Bestellungen"]

def get_price(name):
    preislist = wb["Preisliste"]
    for row in range(2, 12):
        if preislist["a" + str(row)].value == name:
            return preislist["b" + str(row)].value
    return 0

#Preise in Tabelle ausfüllen
for row in range (2, 23):
    product_name = ws["b" + str(row)].value
    ws["d" + str(row)].value = get_price(product_name) # Werte von D-Reihe

#Preise für Bestllungen berechnen
for row in range(2,23):
    c = ws["c" + str(row)].value # Werte von C-Reihe
    d = ws["d" + str(row)].value # Werte von D-Reihe
    result = c * d
    ws["e" + str(row)].value = result

wb.save("Bestellungen.xlsx")