from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "Geburtstage"


ws.append(["Name", "Geburtsdatum", "Spitzname", "Stadt"])

ws["A1"].value = "Name"
ws["B1"].value = "Geburtsdatum"

ws["A2"].value = "Junus"
ws["B2"].value = "03.10.1991"

ws["A3"].value = "Hans Müller"
ws["B3"].value = "03.10.1993"

ws.insert_rows(1)

wb.save("Geburtstage.xlsx")